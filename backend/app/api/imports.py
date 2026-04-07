from io import BytesIO
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List
from app.database import get_db
from app.auth import get_current_user
from app.models.user import User
from app.models.stage import Supplier, Material, Product, ProductBOMItem

router = APIRouter()


def check_admin(current_user: User):
    if current_user.role not in ["SUPER_ADMIN", "COMPANY_ADMIN"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo administradores pueden importar datos",
        )


def _style_header(ws, headers: List[str]):
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 6, 18)


def _style_example_row(ws, row: int, values: list):
    example_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    example_font = Font(color="94A3B8", italic=True)
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = example_fill
        cell.font = example_font


# ---------------------------------------------------------------------------
# GET /import/template — genera y descarga la plantilla Excel
# ---------------------------------------------------------------------------
@router.get("/template")
def download_template(
    current_user: User = Depends(get_current_user),
):
    wb = Workbook()

    # --- Hoja 1: Proveedores ---
    ws1 = wb.active
    ws1.title = "Proveedores"
    headers1 = ["nombre *", "contacto", "email", "telefono", "direccion"]
    _style_header(ws1, headers1)
    _style_example_row(ws1, 2, ["Aceros del Norte S.A.", "Carlos Mendoza", "ventas@aceros.com", "81-2345-6789", "Av. Industrial 450, Monterrey"])
    _style_example_row(ws1, 3, ["Ferretería Central", "Miguel Vargas", "", "81-4567-8901", ""])

    # --- Hoja 2: Materiales ---
    ws2 = wb.create_sheet("Materiales")
    headers2 = ["nombre *", "unidad *", "proveedor"]
    _style_header(ws2, headers2)
    _style_example_row(ws2, 2, ['Lámina de acero 1/8"', "m²", "Aceros del Norte S.A."])
    _style_example_row(ws2, 3, ["Tornillos M8 x 30mm", "unidad", "Ferretería Central"])

    # --- Hoja 3: Productos ---
    ws3 = wb.create_sheet("Productos")
    headers3 = ["nombre *", "sku", "material *", "cantidad_por_unidad *"]
    _style_header(ws3, headers3)
    # Ejemplo: producto con 2 materiales (misma fila de nombre para cada material)
    _style_example_row(ws3, 2, ["Puerta corrediza", "PUERTA-001", 'Lámina de acero 1/8"', "8"])
    _style_example_row(ws3, 3, ["Puerta corrediza", "PUERTA-001", "Tornillos M8 x 30mm", "40"])
    _style_example_row(ws3, 4, ["Mesa industrial", "", 'Lámina de acero 1/8"', "4"])

    # Nota en hoja de productos
    ws3.cell(row=6, column=1, value="NOTA: Un producto con varios materiales ocupa múltiples filas con el mismo nombre.")
    ws3.cell(row=6, column=1).font = Font(color="64748B", italic=True, size=9)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_flowbit.xlsx"},
    )


# ---------------------------------------------------------------------------
# Lógica compartida de parseo del Excel
# ---------------------------------------------------------------------------
def _parse_excel(content: bytes, company_id: int, db: Session):
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)

    # Proveedores existentes en BD
    existing_suppliers = {
        s.name.strip().lower(): s
        for s in db.query(Supplier).filter(Supplier.company_id == company_id).all()
    }

    # Materiales existentes en BD
    existing_materials = {
        m.name.strip().lower(): m
        for m in db.query(Material).filter(Material.company_id == company_id).all()
    }

    # Productos existentes en BD
    existing_products = {
        p.name.strip().lower(): p
        for p in db.query(Product).filter(Product.company_id == company_id).all()
    }

    suppliers_data = []
    materials_data = []
    products_data = []

    # --- Parsear Proveedores ---
    if "Proveedores" in wb.sheetnames:
        ws = wb["Proveedores"]
        seen_in_file = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nombre = row[0]
            if not nombre or str(nombre).strip() == "":
                continue
            nombre = str(nombre).strip()
            nombre_lower = nombre.lower()

            contacto = str(row[1]).strip() if row[1] else None
            email = str(row[2]).strip() if row[2] else None
            telefono = str(row[3]).strip() if row[3] else None
            direccion = str(row[4]).strip() if row[4] else None

            if nombre_lower in existing_suppliers:
                status = "exists"
                message = "Ya existe, se omitirá"
            elif nombre_lower in seen_in_file:
                status = "error"
                message = "Nombre duplicado en el archivo"
            else:
                seen_in_file[nombre_lower] = nombre
                status = "new"
                message = ""

            suppliers_data.append({
                "row": i,
                "nombre": nombre,
                "contacto": contacto,
                "email": email,
                "telefono": telefono,
                "direccion": direccion,
                "status": status,
                "message": message,
            })

    # Combinar para lookup de materiales
    all_supplier_names = {**{k: v.name for k, v in existing_suppliers.items()}}
    for s in suppliers_data:
        if s["status"] == "new":
            all_supplier_names[s["nombre"].lower()] = s["nombre"]

    # --- Parsear Materiales ---
    if "Materiales" in wb.sheetnames:
        ws = wb["Materiales"]
        seen_in_file = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nombre = row[0]
            unidad = row[1]

            if not nombre or str(nombre).strip() == "":
                continue
            nombre = str(nombre).strip()
            unidad = str(unidad).strip() if unidad else ""

            if not unidad:
                materials_data.append({
                    "row": i,
                    "nombre": nombre,
                    "unidad": "",
                    "proveedor": None,
                    "status": "error",
                    "message": "Falta la unidad (obligatoria)",
                })
                continue

            nombre_lower = nombre.lower()
            proveedor_nombre = str(row[2]).strip() if row[2] else None

            if nombre_lower in existing_materials:
                status = "exists"
                message = "Ya existe, se omitirá"
            elif nombre_lower in seen_in_file:
                status = "error"
                message = "Nombre duplicado en el archivo"
            else:
                seen_in_file[nombre_lower] = nombre
                status = "new"
                message = ""

            # Validar proveedor si se especificó
            if proveedor_nombre and proveedor_nombre.lower() not in all_supplier_names:
                status = "error"
                message = f"Proveedor '{proveedor_nombre}' no encontrado"

            materials_data.append({
                "row": i,
                "nombre": nombre,
                "unidad": unidad,
                "proveedor": proveedor_nombre,
                "status": status,
                "message": message,
            })

    # Combinar para lookup de productos
    all_material_names = {**{k: v.name for k, v in existing_materials.items()}}
    for m in materials_data:
        if m["status"] == "new":
            all_material_names[m["nombre"].lower()] = m["nombre"]

    # --- Parsear Productos ---
    if "Productos" in wb.sheetnames:
        ws = wb["Productos"]
        products_map = {}  # nombre_lower -> dict

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nombre = row[0]
            if not nombre or str(nombre).strip() == "" or str(nombre).startswith("NOTA"):
                continue

            nombre = str(nombre).strip()
            nombre_lower = nombre.lower()
            sku = str(row[1]).strip() if row[1] else None
            material_nombre = str(row[2]).strip() if row[2] else None
            cantidad = row[3]

            if nombre_lower not in products_map:
                if nombre_lower in existing_products:
                    prod_status = "exists"
                    prod_message = "Ya existe, se omitirá"
                else:
                    prod_status = "new"
                    prod_message = ""

                products_map[nombre_lower] = {
                    "nombre": nombre,
                    "sku": sku,
                    "status": prod_status,
                    "message": prod_message,
                    "bom_items": [],
                }

            prod = products_map[nombre_lower]
            if prod["status"] == "exists":
                continue

            if not material_nombre:
                prod["status"] = "error"
                prod["message"] = f"Fila {i}: falta el nombre del material"
                continue

            if not cantidad:
                prod["status"] = "error"
                prod["message"] = f"Fila {i}: falta la cantidad para '{material_nombre}'"
                continue

            try:
                qty = float(str(cantidad).replace(",", "."))
                if qty <= 0:
                    raise ValueError
            except ValueError:
                prod["status"] = "error"
                prod["message"] = f"Fila {i}: cantidad inválida '{cantidad}'"
                continue

            if material_nombre.lower() not in all_material_names:
                prod["status"] = "error"
                prod["message"] = f"Material '{material_nombre}' no encontrado"
                continue

            prod["bom_items"].append({
                "material": material_nombre,
                "cantidad": qty,
            })

        products_data = list(products_map.values())
        # Productos sin materiales
        for p in products_data:
            if p["status"] == "new" and len(p["bom_items"]) == 0:
                p["status"] = "error"
                p["message"] = "No se especificaron materiales"

    return {
        "suppliers": suppliers_data,
        "materials": materials_data,
        "products": products_data,
        "_meta": {
            "existing_suppliers": existing_suppliers,
            "existing_materials": existing_materials,
            "all_supplier_names_map": all_supplier_names,
            "all_material_names_map": all_material_names,
        },
    }


# ---------------------------------------------------------------------------
# POST /import/preview — previsualiza sin crear nada
# ---------------------------------------------------------------------------
@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_admin(current_user)
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="Usuario sin empresa")

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    content = await file.read()
    result = _parse_excel(content, current_user.company_id, db)

    return {
        "suppliers": result["suppliers"],
        "materials": result["materials"],
        "products": result["products"],
    }


# ---------------------------------------------------------------------------
# POST /import/confirm — crea los registros en BD
# ---------------------------------------------------------------------------
@router.post("/confirm")
async def confirm_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_admin(current_user)
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="Usuario sin empresa")

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    content = await file.read()
    parsed = _parse_excel(content, current_user.company_id, db)
    company_id = current_user.company_id

    existing_suppliers = parsed["_meta"]["existing_suppliers"]
    existing_materials = parsed["_meta"]["existing_materials"]
    all_supplier_names_map = parsed["_meta"]["all_supplier_names_map"]
    all_material_names_map = parsed["_meta"]["all_material_names_map"]

    created = {"suppliers": 0, "materials": 0, "products": 0}
    skipped = {"suppliers": 0, "materials": 0, "products": 0}

    # Mapa de nombre_lower -> objeto Supplier recién creado
    new_supplier_objects: dict[str, Supplier] = {}
    new_material_objects: dict[str, Material] = {}

    # --- Crear proveedores ---
    for s in parsed["suppliers"]:
        if s["status"] != "new":
            skipped["suppliers"] += 1
            continue
        obj = Supplier(
            company_id=company_id,
            name=s["nombre"],
            contact_name=s["contacto"],
            email=s["email"] if s["email"] else None,
            phone=s["telefono"],
            address=s["direccion"],
        )
        db.add(obj)
        db.flush()
        new_supplier_objects[s["nombre"].lower()] = obj
        created["suppliers"] += 1

    # --- Crear materiales ---
    for m in parsed["materials"]:
        if m["status"] != "new":
            skipped["materials"] += 1
            continue

        supplier_id = None
        if m["proveedor"]:
            proveedor_lower = m["proveedor"].lower()
            if proveedor_lower in new_supplier_objects:
                supplier_id = new_supplier_objects[proveedor_lower].id
            elif proveedor_lower in existing_suppliers:
                supplier_id = existing_suppliers[proveedor_lower].id

        obj = Material(
            company_id=company_id,
            name=m["nombre"],
            unit=m["unidad"],
            supplier_id=supplier_id,
            qty_available=0,
        )
        db.add(obj)
        db.flush()
        new_material_objects[m["nombre"].lower()] = obj
        created["materials"] += 1

    # --- Crear productos ---
    for p in parsed["products"]:
        if p["status"] != "new":
            skipped["products"] += 1
            continue

        prod_obj = Product(
            company_id=company_id,
            name=p["nombre"],
            sku=p["sku"],
        )
        db.add(prod_obj)
        db.flush()

        for bom in p["bom_items"]:
            mat_lower = bom["material"].lower()
            if mat_lower in new_material_objects:
                mat_id = new_material_objects[mat_lower].id
            elif mat_lower in existing_materials:
                mat_id = existing_materials[mat_lower].id
            else:
                continue

            bom_item = ProductBOMItem(
                product_id=prod_obj.id,
                material_id=mat_id,
                qty_per_unit=bom["cantidad"],
            )
            db.add(bom_item)

        created["products"] += 1

    db.commit()

    return {
        "success": True,
        "created": created,
        "skipped": skipped,
    }
